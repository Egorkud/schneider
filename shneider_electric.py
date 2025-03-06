import time
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

from base import Base
from base import ParserLogger


# It would be possibly better to use Selenium everywhere here
class ShneiderElectric(Base):
    def __init__(self):
        super().__init__()
        self.init_project()
        options = webdriver.ChromeOptions()
        service = Service(ChromeDriverManager().install())

        self.driver = webdriver.Chrome(service=service, options=options)

    def scrap(self, filename_new: str = "parsed_data.xlsx",
              data_file: str = "parsing.xlsx"):

        work_sheet = openpyxl.open(data_file).active
        logger = ParserLogger()

        # Row for new file, count for old file: collecting articules
        for row in range(2, work_sheet.max_row + 1):
            print(self.GREEN(f"{row}. Started"))

            item_articule = work_sheet.cell(row, 2).value
            url = work_sheet.cell(row, 3).value
            searched_item_link = self.get_searched_item_link(url)
            print(url)

            self.log_data = [row - 1, item_articule, searched_item_link]

            if not searched_item_link:
                print(self.GREEN(f"Link skipped. None response"))
                logger.log_parsing_result(self.log_data)
                continue

            self.blank_sheet.cell(row, 1).value = url
            self.blank_sheet.cell(row, 2).value = item_articule

            # If there is necessary to scrap UKR version, it is important for names and descriptions

            """Обираємо необхідний парсер і реалізуємо його роботу в окремому файлі"""
            # Uncomment necessary scrappers
            self.get_name(row)
            self.get_descriptions(row)
            self.get_characteristics(row)

            logger.log_parsing_result(self.log_data)

            self.blank_file.save(filename_new)
            time.sleep(1 + random.uniform(0, 1))

        self.driver.quit()
        print(self.GREEN(f"\nFile {filename_new} created"))
        print(self.GREEN(f"Total photo count: {self.images_counter}"))
        print(self.GREEN(f"Total descriptions count: {self.instructions_counter}"))

    def get_searched_item_link(self, url):
        """Шукає товар за артикулом та повертає його URL"""
        try:
            self.driver.get(url)
            time.sleep(2)

            # Отримуємо поточний URL після переходу
            current_url = self.driver.current_url
            if current_url == "https://www.se.com/ua/uk/all-products/":
                return False  # Якщо URL змінився на головну, товарної сторінки немає

            # # Перевіряємо наявність важливого елемента (наприклад, назва товару)
            # self.driver.find_element(By.CLASS_NAME, "main-product-info")

            return True  # Якщо елемент знайдено, це сторінка товару

        except Exception as ex:
            self.log_data.append(f"Error getting search link: {ex}")
            return None

    def get_characteristics(self, row):
        """Отримує характеристики товару"""
        try:
            chars_data = self.driver.execute_script("""
            let root1 = document.querySelector('pes-description-and-specifications');
            if (!root1) return 'Root 1 not found';

            let shadow1 = root1.shadowRoot;
            let root2 = shadow1.querySelector('pes-specifications');
            if (!root2) return 'Root 2 not found';

            let shadow2 = root2.shadowRoot;
            let tables = shadow2.querySelectorAll('pes-specifications-table'); // Отримуємо ВСІ таблиці
            if (!tables.length) return 'Tables not found';

            let result = [];

            tables.forEach(table => {
                let shadowTable = table.shadowRoot;
                if (shadowTable) {
                    let rows = shadowTable.querySelectorAll('.specifications-table__row');
                    rows.forEach(row => {
                        let key = row.querySelector('.specifications-table__row-heading')?.innerText.trim();
                        let value = row.querySelector('.specifications-table__cell')?.innerText.trim();
                        if (key && value) result.push({ [key]: value });
                    });
                } else {
                    result.push('ShadowRoot not accessible');
                }
            });

            return result;
            """)

            for char in chars_data:
                for key, value in char.items():
                    self.check_key(key)

                    # Adding characteristics
                    char_col = self.char_dict[key]
                    self.blank_sheet.cell(row, char_col).value = value

        except Exception as ex:
            self.log_data.append(f"Error getting characteristics: {ex}")

    def get_descriptions(self, row):
        # Get description RU (But now it is better to generate with gpt)
        # OPTIONAL: Get description UKR if it is possible on website (other option is translator)
        try:
            description = self.driver.execute_script("""
            let root1 = document.querySelector('pes-description-and-specifications');
            if (!root1) return 'Root 1 not found';

            let shadow1 = root1.shadowRoot;
            let root2 = shadow1.querySelector('pes-product-description');
            if (!root2) return 'Root 2 not found';

            let shadow2 = root2.shadowRoot;
            let root3 = shadow2.querySelector('pes-description');
            if (!root3) return 'Root 3 not found';

            let shadow3 = root3.shadowRoot;
            if (!shadow3) return 'Shadow 3 not accessible';

            let containers = shadow3.querySelectorAll('div.text-container');
            if (!containers.length) return 'Containers not found';

            for (let container of containers) {
                let description = container.querySelector('.description__content.font-m');
                if (description && description.textContent.trim().length > 0) {
                    return description.textContent.trim();
                }
            }
            return 'Description not found';
            """)

            if description in ("Description not found", "Shadow 3 not accessible", "Root 2 not found"):
                self.log_data.append("Error getting descriptions")
                raise Exception(f"Description not found: {description}")

            self.blank_sheet.cell(row, 12).value = str(description)
        except Exception as ex:
            self.log_data.append(f"Error getting descriptions: {ex}")

    def get_name(self, row):
        # Get names ru ukr
        try:
            # Implement scrapping
            shadow_host = self.driver.find_element(By.CSS_SELECTOR, "pes-main-product-info")
            shadow_root = self.driver.execute_script("return arguments[0].shadowRoot", shadow_host)
            name = shadow_root.find_element(By.CSS_SELECTOR, ".main-product-info__description").text

            self.blank_sheet.cell(row, 5).value = name
        except Exception as ex:
            self.log_data.append(f"Error getting product name {ex}")


if __name__ == '__main__':
    ShneiderElectric().scrap()