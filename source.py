import openpyxl
from openpyxl import Workbook
from pathlib import Path
import conf
from colorama import Fore, Back, Style, init


class Source:
    def __init__(self):
        # Common data (usually does not need changes)
        try:
            self.blank_file = openpyxl.open("data/sample.xlsx")
            self.blank_sheet = self.blank_file.active
            self.book_empty = openpyxl.Workbook()  # Empty table
            self.empty_sheet = self.book_empty.active
        except FileNotFoundError as ex:
            print(ex)
            print("Problems with common data files load. Use init_project()\n")


        # Adding colours for cosy prints
        init(autoreset=True)
        self.GREEN = lambda text: f"{Fore.GREEN}{text}{Style.RESET_ALL}"
        self.RED = lambda text: f"{Fore.RED}{text}{Style.RESET_ALL}"
        self.YELLOW = lambda text: f"{Fore.YELLOW}{text}{Style.RESET_ALL}"
        self.BLUE = lambda text: f"{Fore.BLUE}{text}{Style.RESET_ALL}"

    def close(self):
        try:
            self.blank_file.close()
            self.book_empty.close()
            self.book_names_data.close()
        except AttributeError as ex:
            print(ex)

    def init_project(self):
        def create_path(*files):
            for i in files:
                if not i.exists():
                    if i.suffix:
                        i.touch(exist_ok=True)
                        print(self.GREEN(f"File '{i}' created"))
                    else:
                        i.mkdir(exist_ok=True)
                        print(self.GREEN(f"Directory '{i}' created"))

        def create_excel_file(file_path, columns):
            """Створює Excel-файл з вказаними заголовками, якщо він не існує."""
            if file_path.exists():
                return

            create_path(file_path.parent, file_path)

            wb = Workbook()
            sheet = wb.active
            sheet.title = "Sheet"

            for col_id, col_name in columns.items():
                sheet.cell(1, col_id).value = col_name

            wb.save(file_path)
            print(self.GREEN(f"File {file_path} was filled"))

        print(self.BLUE("\nProject initialisation started\n"))
        # Create data directory and files inside
        data_dir = Path("data")
        sample_file = data_dir / "sample.xlsx"
        create_excel_file(sample_file, conf.SAMPLE_PRODUCT_COLUMNS)

        print(self.BLUE("\nProject initialisation finished\n"))